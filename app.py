from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), 'gastos.db')

# ── BD ──────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        type      TEXT    NOT NULL CHECK(type IN ('income','expense')),
        amount    REAL    NOT NULL,
        category  TEXT    NOT NULL,
        description TEXT,
        date      TEXT    NOT NULL,
        created_at TEXT   DEFAULT (datetime('now'))
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS categories (
        id    INTEGER PRIMARY KEY AUTOINCREMENT,
        name  TEXT UNIQUE NOT NULL,
        type  TEXT NOT NULL CHECK(type IN ('income','expense','both')),
        icon  TEXT DEFAULT '📁',
        color TEXT DEFAULT '#6366f1'
    )''')
    # seed categories
    defaults = [
        ('Salario','income','💼','#10b981'),
        ('Freelance','income','💻','#3b82f6'),
        ('Inversiones','income','📈','#8b5cf6'),
        ('Otros ingresos','income','💰','#f59e0b'),
        ('Alimentación','expense','🍔','#ef4444'),
        ('Transporte','expense','🚗','#f97316'),
        ('Vivienda','expense','🏠','#6366f1'),
        ('Salud','expense','⚕️','#ec4899'),
        ('Educación','expense','📚','#14b8a6'),
        ('Entretenimiento','expense','🎬','#a855f7'),
        ('Ropa','expense','👕','#f43f5e'),
        ('Servicios','expense','⚡','#0ea5e9'),
        ('Tecnología','expense','📱','#8b5cf6'),
        ('Deporte','expense','🏋️','#22c55e'),
        ('Viajes','expense','✈️','#f59e0b'),
        ('Otros gastos','expense','📦','#64748b'),
    ]
    c.executemany('INSERT OR IGNORE INTO categories (name,type,icon,color) VALUES (?,?,?,?)', defaults)
    conn.commit()
    conn.close()

# ── RUTAS PRINCIPALES ────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# ── API TRANSACCIONES ────────────────────────────────────────────────────────
@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    conn = get_db()
    filters, params = [], []
    if request.args.get('type'):
        filters.append('type = ?'); params.append(request.args['type'])
    if request.args.get('category'):
        filters.append('category = ?'); params.append(request.args['category'])
    if request.args.get('date_from'):
        filters.append('date >= ?'); params.append(request.args['date_from'])
    if request.args.get('date_to'):
        filters.append('date <= ?'); params.append(request.args['date_to'])
    if request.args.get('search'):
        filters.append('(description LIKE ? OR category LIKE ?)');
        params += [f'%{request.args["search"]}%'] * 2
    where = ('WHERE ' + ' AND '.join(filters)) if filters else ''
    rows = conn.execute(f'SELECT * FROM transactions {where} ORDER BY date DESC, id DESC', params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    d = request.get_json()
    conn = get_db()
    cur = conn.execute(
        'INSERT INTO transactions (type,amount,category,description,date) VALUES (?,?,?,?,?)',
        (d['type'], float(d['amount']), d['category'], d.get('description',''), d['date'])
    )
    conn.commit()
    row = conn.execute('SELECT * FROM transactions WHERE id=?', (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/transactions/<int:tid>', methods=['PUT'])
def update_transaction(tid):
    d = request.get_json()
    conn = get_db()
    conn.execute(
        'UPDATE transactions SET type=?,amount=?,category=?,description=?,date=? WHERE id=?',
        (d['type'], float(d['amount']), d['category'], d.get('description',''), d['date'], tid)
    )
    conn.commit()
    row = conn.execute('SELECT * FROM transactions WHERE id=?', (tid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
def delete_transaction(tid):
    conn = get_db()
    conn.execute('DELETE FROM transactions WHERE id=?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── API CATEGORÍAS ───────────────────────────────────────────────────────────
@app.route('/api/categories', methods=['GET'])
def get_categories():
    conn = get_db()
    rows = conn.execute('SELECT * FROM categories ORDER BY type, name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/categories', methods=['POST'])
def add_category():
    d = request.get_json()
    conn = get_db()
    try:
        cur = conn.execute('INSERT INTO categories (name,type,icon,color) VALUES (?,?,?,?)',
                           (d['name'], d['type'], d.get('icon','📁'), d.get('color','#6366f1')))
        conn.commit()
        row = conn.execute('SELECT * FROM categories WHERE id=?', (cur.lastrowid,)).fetchone()
        conn.close()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Categoría ya existe'}), 400

@app.route('/api/categories/<int:cid>', methods=['DELETE'])
def delete_category(cid):
    conn = get_db()
    conn.execute('DELETE FROM categories WHERE id=?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── API ESTADÍSTICAS ─────────────────────────────────────────────────────────
@app.route('/api/stats/summary')
def stats_summary():
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    rows = conn.execute(
        "SELECT type, SUM(amount) as total FROM transactions WHERE strftime('%Y-%m',date)=? GROUP BY type",
        (month,)).fetchall()
    conn.close()
    data = {r['type']: r['total'] for r in rows}
    income  = data.get('income', 0)
    expense = data.get('expense', 0)
    return jsonify({'income': income, 'expense': expense, 'balance': income - expense, 'month': month})

@app.route('/api/stats/by_category')
def stats_by_category():
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    ttype = request.args.get('type', 'expense')
    rows = conn.execute(
        "SELECT category, SUM(amount) as total FROM transactions "
        "WHERE type=? AND strftime('%Y-%m',date)=? GROUP BY category ORDER BY total DESC",
        (ttype, month)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/stats/monthly_trend')
def monthly_trend():
    conn = get_db()
    rows = conn.execute(
        "SELECT strftime('%Y-%m',date) as month, type, SUM(amount) as total "
        "FROM transactions GROUP BY month, type ORDER BY month"
    ).fetchall()
    conn.close()
    months = sorted(set(r['month'] for r in rows))[-12:]
    income  = {m: 0 for m in months}
    expense = {m: 0 for m in months}
    for r in rows:
        if r['month'] in months:
            if r['type'] == 'income':  income[r['month']]  = r['total']
            else:                       expense[r['month']] = r['total']
    return jsonify({'months': months, 'income': [income[m] for m in months], 'expense': [expense[m] for m in months]})

@app.route('/api/stats/daily_trend')
def daily_trend():
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    rows = conn.execute(
        "SELECT date, type, SUM(amount) as total FROM transactions "
        "WHERE strftime('%Y-%m',date)=? GROUP BY date, type ORDER BY date",
        (month,)).fetchall()
    conn.close()
    data = defaultdict(lambda: {'income': 0, 'expense': 0})
    for r in rows:
        data[r['date']][r['type']] = r['total']
    dates = sorted(data.keys())
    return jsonify({'dates': dates, 'income': [data[d]['income'] for d in dates], 'expense': [data[d]['expense'] for d in dates]})

# ── EXPORTAR EXCEL ───────────────────────────────────────────────────────────
@app.route('/api/export/excel')
def export_excel():
    conn = get_db()
    rows = conn.execute('SELECT * FROM transactions ORDER BY date DESC').fetchall()
    cats = conn.execute('SELECT * FROM categories').fetchall()
    conn.close()

    wb = openpyxl.Workbook()

    # ── hoja transacciones ──
    ws = wb.active
    ws.title = 'Transacciones'
    header_fill = PatternFill('solid', fgColor='1E293B')
    income_fill  = PatternFill('solid', fgColor='D1FAE5')
    expense_fill = PatternFill('solid', fgColor='FEE2E2')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    headers = ['#','Tipo','Fecha','Categoría','Descripción','Monto (S/)']
    col_widths = [5, 12, 14, 20, 35, 15]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    total_income = total_expense = 0
    for i, r in enumerate(rows, 2):
        fill = income_fill if r['type'] == 'income' else expense_fill
        vals = [r['id'], 'Ingreso' if r['type']=='income' else 'Gasto',
                r['date'], r['category'], r['description'] or '', r['amount']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical='center')
            if col == 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        if r['type'] == 'income':  total_income  += r['amount']
        else:                       total_expense += r['amount']

    last = len(rows) + 2
    ws.cell(row=last, column=5, value='TOTAL INGRESOS').font = Font(bold=True, color='065F46')
    ws.cell(row=last, column=6, value=total_income).number_format = '#,##0.00'
    ws.cell(row=last+1, column=5, value='TOTAL GASTOS').font = Font(bold=True, color='991B1B')
    ws.cell(row=last+1, column=6, value=total_expense).number_format = '#,##0.00'
    ws.cell(row=last+2, column=5, value='BALANCE').font = Font(bold=True)
    bal_cell = ws.cell(row=last+2, column=6, value=total_income - total_expense)
    bal_cell.number_format = '#,##0.00'
    bal_cell.font = Font(bold=True, color='1D4ED8')
    ws.freeze_panes = 'A2'

    # ── hoja resumen mensual ──
    ws2 = wb.create_sheet('Resumen Mensual')
    conn2 = get_db()
    monthly = conn2.execute(
        "SELECT strftime('%Y-%m',date) as month, type, SUM(amount) as total "
        "FROM transactions GROUP BY month,type ORDER BY month"
    ).fetchall()
    conn2.close()
    months_data = defaultdict(lambda: {'income':0,'expense':0})
    for r in monthly:
        months_data[r['month']][r['type']] = r['total']

    for col, h in enumerate(['Mes','Ingresos (S/)','Gastos (S/)','Balance (S/)'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
        ws2.column_dimensions[get_column_letter(col)].width = 18

    for i, m in enumerate(sorted(months_data.keys()), 2):
        inc = months_data[m]['income']
        exp = months_data[m]['expense']
        for col, v in enumerate([m, inc, exp, inc-exp], 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = thin
            if col > 1: cell.number_format = '#,##0.00'

    # ── hoja categorías ──
    ws3 = wb.create_sheet('Categorías')
    for col, h in enumerate(['ID','Nombre','Tipo','Ícono'], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[get_column_letter(col)].width = 20
    for i, c in enumerate(cats, 2):
        ws3.cell(row=i, column=1, value=c['id'])
        ws3.cell(row=i, column=2, value=c['name'])
        ws3.cell(row=i, column=3, value='Ingreso' if c['type']=='income' else 'Gasto' if c['type']=='expense' else 'Ambos')
        ws3.cell(row=i, column=4, value=c['icon'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"gastos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
